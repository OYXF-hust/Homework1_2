import xlrd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from xlrd import open_workbook
from tkinter import *
from tkinter import ttk
import os


def picture_make():
    s = s_input.get()
    if s == '条形图':
        def read_excel():
            workbook = xlrd.open_workbook(os.path.join(os.getcwd(), 'data1.xlsx'))
            sheet = workbook.sheet_by_name("Sheet1")
            row_num = sheet.nrows
            row_list = []
            for i in range(0, row_num):
                row_data = sheet.row_values(i)
                for data in row_data:
                    if data != '':
                        row_list.append(data)

            return sorted(row_list)


        def draw_picture(row_list):

            plt.ylabel("speed of Pedestrians 行人速度", fontproperties='SimHei')
            plt.bar(range(len(row_list)), row_list)
            plt.savefig("D:\description/条形图.png")
            #plt.show()
        if __name__ == '__main__':
            draw_picture(read_excel())
    if s == '折线图':
        x_data1 = []
        y_data1 = []
        wb = open_workbook('data.xlsx')
        for s in wb.sheets():
            for row in range(s.nrows):
                values = []
                for col in range(s.ncols):
                    values.append(s.cell(row, col).value)
                x_data1.append(values[0])
                y_data1.append(values[1])
        plt.plot(x_data1, y_data1, 'bo-', label=u"Phase curve", linewidth=1)
        plt.title(u"Line chart")
        plt.legend()
        plt.xlabel(u"input-deg")
        plt.ylabel(u"output-V")
        plt.savefig("D:\description/折线图.png")
        #plt.show()
    if s == '散点图':
        x_data1 = []
        y_data1 = []
        wb = open_workbook('data.xlsx')
        for s in wb.sheets():
            for row in range(s.nrows):
                values = []
                for col in range(s.ncols):
                    values.append(s.cell(row, col).value)
                x_data1.append(values[0])
                y_data1.append(values[1])
        plt.scatter(x_data1, y_data1, label=u"Piont")
        plt.title(u"Scatter diagram")
        plt.legend()
        plt.xlabel(u"input-deg")
        plt.ylabel(u"output-V")
        plt.savefig("D:\description/散点图.png")
        #plt.show()
    if s == '饼图':
        book = load_workbook(filename=r"data2.xlsx")
        sheet = book["Sheet1"]
        edu = []
        labels = []
        row_num = 1
        while row_num <= 5:
            labels.append(sheet.cell(row=row_num, column=1).value)
            row_num = row_num + 1
        row_num = 1
        while row_num <= 5:
            edu.append(sheet.cell(row=row_num, column=2).value)
            row_num = row_num + 1
        plt.style.use('ggplot')
        colors = ['#9999ff', '#ff9999', '#7777aa', '#2442aa', '#dd5555']
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False
        plt.axes(aspect='equal')
        plt.xlim(0, 4)
        plt.ylim(0, 4)
        plt.pie(x=edu,
                labels=labels,
                colors=colors,
                autopct='%.1f%%',
                pctdistance=0.8,
                labeldistance=1.15,
                startangle=180,
                radius=1.5,
                counterclock=False,
                wedgeprops={'linewidth': 1.5, 'edgecolor': 'green'},
                textprops={'fontsize': 12, 'color': 'k'},
                center=(1.8, 1.8),
                frame=1)
        plt.xticks(())
        plt.yticks(())
        plt.title('Sector diagram')
        plt.savefig("D:\description/饼图.png")
        #plt.show()
    text.insert(END, "制图成功")
    text.see(END)
    text.update()


def main():
    global s_input, text
    root = Tk()
    root.title('Drawing')
    root.geometry('280x325')
    Label(root, text='请选择图表类型:', font=("华文行楷", 15), fg='black').grid()

    s_input = StringVar()
    numberChosen = ttk.Combobox(root, width=25, textvariable=s_input)
    numberChosen['values'] = ("折线图", "条形图", "饼图", "散点图")  # 设置下拉列表的值
    numberChosen.grid(column=0, row=1)  # 设置其在界面中出现的位置  column代表列   row 代表行
    numberChosen.current(0)  # 设置下拉列表默认显示的值，0为 numberChosen['values'] 的下标值
    Label(root, text='运行结果:', font=("宋体", 15), fg='black').grid(row=4, columnspan=1)
    text = Listbox(root, font=('微软雅黑', 15), width=23, height=5)
    text.grid(row=5, columnspan=1)
    button = Button(root, text='制图', font=("微软雅黑", 15), command=picture_make).grid(row=6, column=0, sticky=W)
    button = Button(root, text='退出', font=("微软雅黑", 15), command=root.quit).grid(row=6, column=0, sticky=E)
    mainloop()


main()